import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF
from datetime import datetime

st.set_page_config(
    page_title="Ressources AWS",
    page_icon="☁️",
    layout="wide",
    initial_sidebar_state="expanded"
)

CSV_PATH = "attached_assets/ec2_prices_20260410-140511_1782061605327.csv"

REGION_FLAGS = {
    "eu-central-1": "🇩🇪",
    "eu-central-2": "🇨🇭",
    "eu-north-1": "🇸🇪",
    "eu-south-1": "🇮🇹",
    "eu-south-2": "🇪🇸",
    "eu-west-1": "🇮🇪",
    "eu-west-2": "🇬🇧",
    "eu-west-3": "🇫🇷",
    "us-east-1": "🇺🇸",
    "us-east-2": "🇺🇸",
}

REGION_CITIES = {
    "eu-central-1": "Frankfurt, Allemagne",
    "eu-central-2": "Zurich, Suisse",
    "eu-north-1": "Stockholm, Suède",
    "eu-south-1": "Milan, Italie",
    "eu-south-2": "Madrid, Espagne",
    "eu-west-1": "Dublin, Irlande",
    "eu-west-2": "Londres, Royaume-Uni",
    "eu-west-3": "Paris, France",
    "us-east-1": "Virginie du Nord, USA",
    "us-east-2": "Ohio, USA",
}

@st.cache_data
def load_data():
    df = pd.read_csv(CSV_PATH, sep=";", decimal=",", low_memory=False)
    numeric_cols = [
        "CostH1yr","CostMRR1yr","CostYRR1yr",
        "CostH3yr","CostMRR3yr","CostYRR3yr",
        "ODH","ODMRR","ODYRR",
        "SnapG/M","gp2G/M","gp3G/M","io1G/M","st1G/M",
        "s3<50T","s3<450T","s3>500T","glacierG/M",
        "CPU","RAM","SAPS"
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    df["Region Code"] = df["Region Code"].str.strip()
    df["Region Name"] = df["Region Name"].str.strip()
    df["Ec2Type"] = df["Ec2Type"].str.strip()
    df["OS"] = df["OS"].str.strip()
    df["SavingsPlanH"] = df["CostH1yr"] * 1.05
    df["SavingsPlanMRR"] = df["CostMRR1yr"] * 1.05
    return df

df = load_data()

with st.sidebar:
    logo_col1, logo_col2 = st.columns(2)
    with logo_col1:
        st.image("https://upload.wikimedia.org/wikipedia/commons/9/93/Amazon_Web_Services_Logo.svg", width=110)
    with logo_col2:
        st.image("attached_assets/oxya_new_logo_CMYK_site-web_1782063768056.png", width=110)
    st.markdown("## Navigation")
    page = st.radio(
        "Choisir une section",
        [
            "🏠 Accueil",
            "💻 Ressources EC2",
            "📊 Analyse coûts EC2",
            "🗄️ Coûts S3 & EBS",
            "🎯 Recommandations",
            "⚖️ Comparaison TCO",
            "🧮 Panier d'instances",
        ],
        label_visibility="collapsed"
    )
    st.markdown("---")
    st.caption("Données : AWS Pricing · Mise à jour : Avril 2026")
    regions_available = sorted(df["Region Name"].unique())
    st.markdown(f"**{len(regions_available)} régions** · **{len(df):,} instances**")


if page == "🏠 Accueil":
    st.title("☁️ Ressources AWS")
    st.markdown("### Plateforme de visualisation des coûts et ressources Amazon Web Services par région")
    st.markdown("---")

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Régions couvertes", len(df["Region Code"].unique()))
    with col2:
        st.metric("Types d'instances EC2", df["Ec2Type"].nunique())
    with col3:
        st.metric("Systèmes d'exploitation", df["OS"].nunique())
    with col4:
        st.metric("Enregistrements totaux", f"{len(df):,}")

    st.markdown("---")
    st.subheader("🌍 Régions disponibles")
    region_summary = df.groupby(["Region Code","Region Name"]).agg(
        instances=("Ec2Type","nunique"),
        od_min=("ODH","min"),
        od_max=("ODH","max"),
    ).reset_index()

    cols = st.columns(3)
    for i, row in region_summary.iterrows():
        flag = REGION_FLAGS.get(row["Region Code"], "🌐")
        city = REGION_CITIES.get(row["Region Code"], row["Region Name"])
        with cols[i % 3]:
            st.markdown(f"""
**{flag} {row['Region Name']}**  
📍 {city}  
🖥️ {row['instances']} types d'instances  
💰 On-Demand: \${row['od_min']:.4f} – \${row['od_max']:.3f} /h
""")
            st.markdown("---")

    st.markdown("---")
    st.subheader("📚 À propos des ressources AWS présentées")

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown("#### 💻 Amazon EC2")
        st.markdown("""
**Elastic Compute Cloud** est le service de calcul à la demande d'AWS.

- Instances virtuelles dans le cloud
- Différents types : General Purpose, Compute Optimized, Memory Optimized
- **Modèles de tarification** : On-Demand, Reserved (1 an, 3 ans), Savings Plans
- Facturation à la seconde (minimum 60s)
- Disponible dans toutes les régions mondiales
        """)

    with c2:
        st.markdown("#### 🗃️ Amazon S3")
        st.markdown("""
**Simple Storage Service** est le service de stockage objet d'AWS.

- Stockage illimité et hautement disponible
- Durabilité de 99,999999999% (11 neuf)
- **Tarification par palier** : < 50 TB, < 450 TB, > 500 TB
- Archivage via Amazon Glacier
- Réplication multi-régions disponible
        """)

    with c3:
        st.markdown("#### 💾 Amazon EBS")
        st.markdown("""
**Elastic Block Store** est le service de stockage en bloc pour EC2.

- Volumes persistants attachés aux instances EC2
- **Types de volumes** : gp2, gp3, io1, st1
- Facturation au Go/mois provisionné
- Snapshots facturés séparément
- Performance ajustable selon le type
        """)

    st.markdown("---")
    st.subheader("📈 Comparaison rapide des coûts On-Demand par région")
    pivot = df.groupby("Region Name")["ODH"].mean().reset_index()
    pivot.columns = ["Région", "Coût horaire moyen ($)"]
    pivot = pivot.sort_values("Coût horaire moyen ($)", ascending=True)

    fig = px.bar(
        pivot,
        x="Coût horaire moyen ($)",
        y="Région",
        orientation="h",
        color="Coût horaire moyen ($)",
        color_continuous_scale="Blues",
        title="Coût horaire moyen On-Demand EC2 par région"
    )
    fig.update_layout(height=400, showlegend=False, coloraxis_showscale=False)
    st.plotly_chart(fig, width="stretch")


elif page == "💻 Ressources EC2":
    st.title("💻 Instances EC2 par Région")
    st.markdown("Explorez et comparez les instances EC2 disponibles dans chaque région AWS.")
    st.markdown("---")

    col_f1, col_f2, col_f3 = st.columns(3)
    with col_f1:
        selected_regions = st.multiselect(
            "Régions",
            options=sorted(df["Region Name"].unique()),
            default=sorted(df["Region Name"].unique())[:3]
        )
    with col_f2:
        selected_os = st.multiselect(
            "Système d'exploitation",
            options=sorted(df["OS"].dropna().unique()),
            default=sorted(df["OS"].dropna().unique())
        )
    with col_f3:
        family_options = sorted(df["Ec2Type"].dropna().str.extract(r'^([a-z0-9]+)\.')[0].dropna().unique())
        selected_family = st.multiselect("Famille d'instances", options=family_options, default=family_options[:5])

    if not selected_regions:
        st.warning("Veuillez sélectionner au moins une région.")
        st.stop()

    fdf = df[df["Region Name"].isin(selected_regions)]
    if selected_os:
        fdf = fdf[fdf["OS"].isin(selected_os)]
    if selected_family:
        fdf = fdf[fdf["Ec2Type"].str.extract(r'^([a-z0-9]+)\.')[0].isin(selected_family)]

    st.markdown(f"**{len(fdf):,} instances** correspondent aux filtres sélectionnés.")
    st.markdown("---")

    st.subheader("🔥 Comparaison des coûts On-Demand par région")
    fig_box = px.box(
        fdf,
        x="Region Name",
        y="ODH",
        color="Region Name",
        title="Distribution des coûts horaires On-Demand par région",
        labels={"ODH": "Coût horaire ($)", "Region Name": "Région"}
    )
    fig_box.update_layout(height=450, showlegend=False, xaxis_tickangle=-30)
    st.plotly_chart(fig_box, width="stretch")

    st.subheader("📊 Coût moyen par famille d'instance et région")
    fdf2 = fdf.copy()
    fdf2["Famille"] = fdf2["Ec2Type"].str.extract(r'^([a-z0-9]+)\.')
    pivot2 = fdf2.groupby(["Famille","Region Name"])["ODH"].mean().reset_index()
    fig_heat = px.density_heatmap(
        pivot2,
        x="Region Name",
        y="Famille",
        z="ODH",
        color_continuous_scale="YlOrRd",
        title="Coût horaire moyen On-Demand ($) par famille & région",
        labels={"ODH": "$/h", "Region Name": "Région", "Famille": "Famille"}
    )
    fig_heat.update_layout(height=500, xaxis_tickangle=-30)
    st.plotly_chart(fig_heat, width="stretch")

    st.subheader("🔍 CPU vs RAM vs Coût")
    fig_scatter = px.scatter(
        fdf,
        x="CPU",
        y="RAM",
        size="ODH",
        color="Region Name",
        hover_data=["Ec2Type","ODH","OS"],
        title="CPU / RAM / Coût On-Demand par région",
        labels={"CPU": "vCPU", "RAM": "RAM (Go)", "ODH": "Coût $/h"}
    )
    fig_scatter.update_layout(height=500)
    st.plotly_chart(fig_scatter, width="stretch")

    st.subheader("📋 Tableau détaillé")
    display_cols = ["Region Name","Ec2Type","OS","CPU","RAM","ODH","ODMRR","CostH1yr","CostH3yr"]
    available_cols = [c for c in display_cols if c in fdf.columns]
    show_df = fdf[available_cols].copy()
    show_df.columns = ["Région","Type","OS","vCPU","RAM (Go)","$/h OD","$/mois OD","$/h 1an","$/h 3ans"][:len(available_cols)]
    st.dataframe(
        show_df.sort_values("$/h OD").head(200),
        width="stretch",
        hide_index=True
    )


elif page == "📊 Analyse coûts EC2":
    st.title("📊 Analyse des Coûts EC2 par Région et Engagement")
    st.markdown("Comparez les coûts EC2 selon le type d'engagement : On-Demand, 1 an, 3 ans et Savings Plans.")
    st.markdown("---")

    c1, c2 = st.columns(2)
    with c1:
        sel_region = st.selectbox(
            "Région principale",
            options=sorted(df["Region Name"].unique()),
            index=0
        )
    with c2:
        compare_regions = st.multiselect(
            "Comparer avec d'autres régions",
            options=[r for r in sorted(df["Region Name"].unique()) if r != sel_region],
            default=[]
        )

    col_f1, col_f2 = st.columns(2)
    with col_f1:
        sel_os = st.selectbox("OS", options=["Tous"] + sorted(df["OS"].dropna().unique().tolist()))
    with col_f2:
        commitment = st.multiselect(
            "Types d'engagement",
            options=["On-Demand (PayAsYouGo)","Reserved 1 an","Reserved 3 ans","Savings Plan (~1 an)"],
            default=["On-Demand (PayAsYouGo)","Reserved 1 an","Reserved 3 ans","Savings Plan (~1 an)"]
        )

    all_regions = [sel_region] + compare_regions
    rdf = df[df["Region Name"].isin(all_regions)].copy()
    if sel_os != "Tous":
        rdf = rdf[rdf["OS"] == sel_os]

    cost_map = {
        "On-Demand (PayAsYouGo)": "ODH",
        "Reserved 1 an": "CostH1yr",
        "Reserved 3 ans": "CostH3yr",
        "Savings Plan (~1 an)": "SavingsPlanH"
    }

    st.markdown("---")
    st.subheader(f"💰 Coûts mensuels comparés par type d'engagement")

    melted_rows = []
    for reg in all_regions:
        reg_df = rdf[rdf["Region Name"] == reg]
        for eng in commitment:
            col_key = cost_map.get(eng)
            if col_key and col_key in reg_df.columns:
                avg_mrr = reg_df[col_key].mean() * 730
                melted_rows.append({"Région": reg, "Engagement": eng, "Coût mensuel moyen ($)": avg_mrr})

    if melted_rows:
        melted_df = pd.DataFrame(melted_rows)
        fig_bar = px.bar(
            melted_df,
            x="Engagement",
            y="Coût mensuel moyen ($)",
            color="Région",
            barmode="group",
            title="Coût mensuel moyen (730h) par type d'engagement et région",
            color_discrete_sequence=px.colors.qualitative.Set2
        )
        fig_bar.update_layout(height=450, xaxis_tickangle=-15)
        st.plotly_chart(fig_bar, width="stretch")

    st.subheader("📉 Économies réalisées vs On-Demand")
    savings_rows = []
    for reg in all_regions:
        reg_df = rdf[rdf["Region Name"] == reg]
        od_avg = reg_df["ODH"].mean()
        if od_avg > 0:
            for eng in ["Reserved 1 an","Reserved 3 ans","Savings Plan (~1 an)"]:
                col_key = cost_map.get(eng)
                if col_key and col_key in reg_df.columns:
                    eng_avg = reg_df[col_key].mean()
                    saving_pct = (1 - eng_avg / od_avg) * 100
                    savings_rows.append({"Région": reg, "Engagement": eng, "Économie (%)": saving_pct})

    if savings_rows:
        sdf = pd.DataFrame(savings_rows)
        fig_sav = px.bar(
            sdf,
            x="Engagement",
            y="Économie (%)",
            color="Région",
            barmode="group",
            title="Économie moyenne vs On-Demand par type d'engagement",
            color_discrete_sequence=px.colors.qualitative.Set2
        )
        fig_sav.update_layout(height=400)
        fig_sav.add_hline(y=0, line_dash="dot", line_color="gray")
        st.plotly_chart(fig_sav, width="stretch")

    st.markdown("---")
    st.subheader(f"🏆 Instances les plus chères par région")
    top_n = st.slider("Nombre d'instances à afficher", 5, 20, 10)

    tabs = st.tabs(all_regions)
    for i, reg in enumerate(all_regions):
        with tabs[i]:
            reg_top = rdf[rdf["Region Name"] == reg].nlargest(top_n, "ODH")[
                ["Ec2Type","OS","CPU","RAM","ODH","ODMRR","ODYRR","CostH1yr","CostH3yr"]
            ].copy()
            reg_top.columns = ["Type","OS","vCPU","RAM(Go)","$/h OD","$/mois OD","$/an OD","$/h 1an","$/h 3ans"]
            st.dataframe(reg_top, width="stretch", hide_index=True)

            fig_top = px.bar(
                reg_top.sort_values("$/h OD", ascending=True),
                x="$/h OD",
                y="Type",
                orientation="h",
                color="$/h OD",
                color_continuous_scale="Reds",
                title=f"Top {top_n} instances les plus chères — {reg}"
            )
            fig_top.update_layout(height=400, coloraxis_showscale=False)
            st.plotly_chart(fig_top, width="stretch")

    st.markdown("---")
    st.subheader("🗓️ Projection de coûts annuels")
    col_p1, col_p2 = st.columns(2)
    with col_p1:
        num_instances = st.number_input("Nombre d'instances", min_value=1, max_value=10000, value=10)
    with col_p2:
        instance_filter = st.text_input("Filtrer par type (ex: m5, c6i)", value="")

    proj_df = rdf[rdf["Region Name"] == sel_region].copy()
    if instance_filter:
        proj_df = proj_df[proj_df["Ec2Type"].str.contains(instance_filter, case=False, na=False)]

    if not proj_df.empty:
        proj_rows = []
        for eng in commitment:
            col_key = cost_map.get(eng)
            if col_key and col_key in proj_df.columns:
                annual = proj_df[col_key].mean() * 8760 * num_instances
                proj_rows.append({"Engagement": eng, "Coût annuel total ($)": annual})

        proj_df2 = pd.DataFrame(proj_rows)
        fig_proj = px.bar(
            proj_df2,
            x="Engagement",
            y="Coût annuel total ($)",
            color="Engagement",
            title=f"Projection annuelle pour {num_instances} instances — {sel_region}",
            color_discrete_sequence=px.colors.qualitative.Pastel
        )
        fig_proj.update_layout(height=400, showlegend=False)
        st.plotly_chart(fig_proj, width="stretch")


elif page == "🗄️ Coûts S3 & EBS":
    st.title("🗄️ Coûts S3 & EBS par Région")
    st.markdown("Visualisez et comparez les coûts de stockage objet (S3) et bloc (EBS) dans chaque région AWS.")
    st.markdown("---")

    region_storage = df.groupby(["Region Code","Region Name"]).agg(
        gp2=("gp2G/M","first"),
        gp3=("gp3G/M","first"),
        io1=("io1G/M","first"),
        st1=("st1G/M","first"),
        snap=("SnapG/M","first"),
        s3_50=("s3<50T","first"),
        s3_450=("s3<450T","first"),
        s3_500=("s3>500T","first"),
        glacier=("glacierG/M","first"),
    ).reset_index()

    st.subheader("💾 EBS — Coûts par type de volume et région")

    ebs_melt = region_storage.melt(
        id_vars=["Region Code","Region Name"],
        value_vars=["gp2","gp3","io1","st1"],
        var_name="Type EBS",
        value_name="$/Go/mois"
    )
    ebs_labels = {"gp2": "GP2 (SSD généraliste)", "gp3": "GP3 (SSD performant)", "io1": "IO1 (SSD haute perf.)", "st1": "ST1 (HDD séquentiel)"}
    ebs_melt["Type EBS"] = ebs_melt["Type EBS"].map(ebs_labels)

    fig_ebs = px.bar(
        ebs_melt.dropna(),
        x="Region Name",
        y="$/Go/mois",
        color="Type EBS",
        barmode="group",
        title="Tarifs EBS par type de volume et région ($/Go/mois)",
        color_discrete_sequence=px.colors.qualitative.Set1,
        labels={"Region Name": "Région"}
    )
    fig_ebs.update_layout(height=450, xaxis_tickangle=-30)
    st.plotly_chart(fig_ebs, width="stretch")

    c1, c2 = st.columns(2)
    with c1:
        st.subheader("📊 Comparaison GP2 vs GP3")
        fig_gp = px.scatter(
            region_storage.dropna(subset=["gp2","gp3"]),
            x="gp2",
            y="gp3",
            text="Region Name",
            title="GP2 vs GP3 — économies potentielles",
            labels={"gp2": "GP2 $/Go/mois", "gp3": "GP3 $/Go/mois"},
            color_discrete_sequence=["#636EFA"]
        )
        max_val = max(region_storage["gp2"].max(), region_storage["gp3"].max()) * 1.1
        fig_gp.add_shape(type="line", x0=0, y0=0, x1=max_val, y1=max_val,
                         line=dict(dash="dash", color="red"), name="Équivalence")
        fig_gp.update_traces(textposition="top center")
        fig_gp.update_layout(height=400)
        st.plotly_chart(fig_gp, width="stretch")
        st.caption("Les points sous la ligne rouge indiquent que GP3 est moins cher que GP2.")

    with c2:
        st.subheader("🧊 Snapshot & Glacier")
        snap_melt = region_storage.melt(
            id_vars=["Region Name"],
            value_vars=["snap","glacier"],
            var_name="Type",
            value_name="$/Go/mois"
        )
        snap_melt["Type"] = snap_melt["Type"].map({"snap": "Snapshot EBS", "glacier": "Glacier (Archivage)"})
        fig_snap = px.bar(
            snap_melt.dropna(),
            x="Region Name",
            y="$/Go/mois",
            color="Type",
            barmode="group",
            title="Coûts Snapshot & Glacier par région",
            labels={"Region Name": "Région"},
            color_discrete_sequence=["#00CC96","#AB63FA"]
        )
        fig_snap.update_layout(height=400, xaxis_tickangle=-30)
        st.plotly_chart(fig_snap, width="stretch")

    st.markdown("---")
    st.subheader("🪣 Amazon S3 — Tarifs par palier de stockage et région")

    s3_melt = region_storage.melt(
        id_vars=["Region Code","Region Name"],
        value_vars=["s3_50","s3_450","s3_500"],
        var_name="Palier S3",
        value_name="$/Go/mois"
    )
    s3_labels = {
        "s3_50": "Tranche 1 : 0–50 TB",
        "s3_450": "Tranche 2 : 50–500 TB",
        "s3_500": "Tranche 3 : > 500 TB"
    }
    s3_melt["Palier S3"] = s3_melt["Palier S3"].map(s3_labels)

    fig_s3 = px.bar(
        s3_melt.dropna(),
        x="Region Name",
        y="$/Go/mois",
        color="Palier S3",
        barmode="group",
        title="Tarifs S3 Standard par palier de stockage et région",
        labels={"Region Name": "Région"},
        color_discrete_sequence=px.colors.qualitative.Pastel1
    )
    fig_s3.update_layout(height=450, xaxis_tickangle=-30)
    st.plotly_chart(fig_s3, width="stretch")

    st.subheader("💡 Simulateur de coût de stockage mensuel")
    col_s1, col_s2, col_s3 = st.columns(3)
    with col_s1:
        sim_region = st.selectbox("Région", options=sorted(df["Region Name"].unique()), key="sim_reg")
        s3_vol = st.number_input("Volume S3 (To)", min_value=0.0, value=10.0, step=1.0)
    with col_s2:
        ebs_gp3_vol = st.number_input("Volume EBS gp3 (Go)", min_value=0, value=500, step=100)
        ebs_io1_vol = st.number_input("Volume EBS io1 (Go)", min_value=0, value=0, step=100)
    with col_s3:
        snap_vol = st.number_input("Snapshots EBS (Go)", min_value=0, value=200, step=50)
        glacier_vol = st.number_input("Glacier (Go)", min_value=0, value=0, step=100)

    sim_row = region_storage[region_storage["Region Name"] == sim_region].iloc[0] if not region_storage[region_storage["Region Name"] == sim_region].empty else None

    if sim_row is not None:
        s3_gb = s3_vol * 1024
        s3_cost = 0.0
        if s3_gb <= 50 * 1024:
            s3_cost = s3_gb * sim_row["s3_50"]
        elif s3_gb <= 500 * 1024:
            s3_cost = (50 * 1024 * sim_row["s3_50"]) + ((s3_gb - 50 * 1024) * sim_row["s3_450"])
        else:
            s3_cost = (50 * 1024 * sim_row["s3_50"]) + (450 * 1024 * sim_row["s3_450"]) + ((s3_gb - 500 * 1024) * sim_row["s3_500"])

        ebs_gp3_cost = ebs_gp3_vol * sim_row["gp3"]
        ebs_io1_cost = ebs_io1_vol * sim_row["io1"]
        snap_cost = snap_vol * sim_row["snap"]
        glacier_cost = glacier_vol * sim_row["glacier"]
        total_cost = s3_cost + ebs_gp3_cost + ebs_io1_cost + snap_cost + glacier_cost

        sim_data = {
            "Composant": ["S3 Standard", "EBS GP3", "EBS IO1", "Snapshots EBS", "Glacier"],
            "Coût mensuel ($)": [s3_cost, ebs_gp3_cost, ebs_io1_cost, snap_cost, glacier_cost]
        }
        sim_df = pd.DataFrame(sim_data)
        sim_df_filtered = sim_df[sim_df["Coût mensuel ($)"] > 0]

        col_r1, col_r2 = st.columns([1, 2])
        with col_r1:
            st.metric("💰 Coût mensuel total estimé", f"${total_cost:,.2f}")
            st.metric("💰 Coût annuel total estimé", f"${total_cost * 12:,.2f}")
            st.markdown("**Détail par composant :**")
            for _, row_s in sim_df.iterrows():
                if row_s["Coût mensuel ($)"] > 0:
                    st.markdown(f"- **{row_s['Composant']}** : ${row_s['Coût mensuel ($)']:,.2f}/mois")

        with col_r2:
            if not sim_df_filtered.empty:
                fig_pie = px.pie(
                    sim_df_filtered,
                    values="Coût mensuel ($)",
                    names="Composant",
                    title=f"Répartition des coûts — {sim_region}",
                    hole=0.4,
                    color_discrete_sequence=px.colors.qualitative.Set2
                )
                fig_pie.update_layout(height=350)
                st.plotly_chart(fig_pie, width="stretch")

    st.markdown("---")
    st.subheader("📋 Tableau récapitulatif des tarifs de stockage")
    display_storage = region_storage[[
        "Region Name","gp2","gp3","io1","st1","snap","s3_50","s3_450","s3_500","glacier"
    ]].copy()
    display_storage.columns = [
        "Région","EBS gp2 ($/Go/m)","EBS gp3 ($/Go/m)","EBS io1 ($/Go/m)","EBS st1 ($/Go/m)",
        "Snapshot ($/Go/m)","S3 <50TB ($/Go/m)","S3 <500TB ($/Go/m)","S3 >500TB ($/Go/m)","Glacier ($/Go/m)"
    ]
    st.dataframe(display_storage, width="stretch", hide_index=True)


elif page == "🎯 Recommandations":
    st.title("🎯 Recommandations Intelligentes")
    st.markdown("Obtenez les meilleures instances EC2 et régions selon votre budget et vos besoins techniques.")
    st.markdown("---")

    st.subheader("⚙️ Définissez vos critères")
    col_a, col_b, col_c = st.columns(3)
    with col_a:
        st.markdown("**💰 Budget**")
        budget_type = st.selectbox("Type de budget", ["Mensuel ($)", "Annuel ($)", "Horaire ($)"], key="reco_budget_type")
        budget_val = st.number_input("Montant", min_value=0.01, value=500.0, step=50.0, key="reco_budget_val")
        engagement = st.selectbox(
            "Type d'engagement",
            ["On-Demand (PayAsYouGo)", "Reserved 1 an", "Reserved 3 ans", "Savings Plan (~1 an)"],
            key="reco_engagement"
        )
    with col_b:
        st.markdown("**🖥️ Ressources**")
        min_cpu = st.number_input("vCPU minimum", min_value=1, max_value=512, value=4, key="reco_cpu")
        min_ram = st.number_input("RAM minimum (Go)", min_value=1, max_value=4096, value=16, key="reco_ram")
        sel_os_reco = st.selectbox("Système d'exploitation", ["Tous"] + sorted(df["OS"].dropna().unique().tolist()), key="reco_os")
    with col_c:
        st.markdown("**🌍 Localisation**")
        pref_regions = st.multiselect(
            "Régions préférées (laisser vide = toutes)",
            options=sorted(df["Region Name"].unique()),
            default=[],
            key="reco_regions"
        )
        max_results = st.slider("Nombre de recommandations", 3, 20, 10, key="reco_max")
        prioritize = st.radio("Optimiser par", ["Meilleur rapport qualité/prix", "Plus économique", "Plus performant (SAPS)"], key="reco_prio")

    cost_col_map = {
        "On-Demand (PayAsYouGo)": "ODH",
        "Reserved 1 an": "CostH1yr",
        "Reserved 3 ans": "CostH3yr",
        "Savings Plan (~1 an)": "SavingsPlanH",
    }
    cost_col = cost_col_map[engagement]

    if budget_type == "Mensuel ($)":
        max_hourly = budget_val / 730
    elif budget_type == "Annuel ($)":
        max_hourly = budget_val / 8760
    else:
        max_hourly = budget_val

    reco_df = df.copy()
    if pref_regions:
        reco_df = reco_df[reco_df["Region Name"].isin(pref_regions)]
    if sel_os_reco != "Tous":
        reco_df = reco_df[reco_df["OS"] == sel_os_reco]
    reco_df = reco_df[reco_df["CPU"] >= min_cpu]
    reco_df = reco_df[reco_df["RAM"] >= min_ram]
    reco_df = reco_df[reco_df[cost_col].notna() & (reco_df[cost_col] <= max_hourly)]

    if reco_df.empty:
        st.warning("⚠️ Aucune instance ne correspond à vos critères. Essayez d'augmenter le budget ou de réduire les exigences CPU/RAM.")
        st.stop()

    reco_df = reco_df.copy()
    reco_df["coût_horaire"] = reco_df[cost_col]
    reco_df["coût_mensuel"] = reco_df[cost_col] * 730
    reco_df["coût_annuel"] = reco_df[cost_col] * 8760
    reco_df["saps_norm"] = reco_df["SAPS"].fillna(0)
    reco_df["score_qp"] = (reco_df["saps_norm"] / (reco_df["coût_horaire"] + 0.001)).fillna(0)
    reco_df["économie_vs_od"] = ((1 - reco_df[cost_col] / reco_df["ODH"]) * 100).round(1)

    if prioritize == "Meilleur rapport qualité/prix":
        reco_df = reco_df.sort_values("score_qp", ascending=False)
    elif prioritize == "Plus économique":
        reco_df = reco_df.sort_values("coût_horaire", ascending=True)
    else:
        reco_df = reco_df.sort_values("saps_norm", ascending=False)

    top_reco = reco_df.head(max_results).reset_index(drop=True)

    st.markdown("---")
    st.subheader(f"🏆 Top {len(top_reco)} recommandations")

    budget_monthly = budget_val if budget_type == "Mensuel ($)" else (budget_val / 12 if budget_type == "Annuel ($)" else budget_val * 730)
    c_info1, c_info2, c_info3, c_info4 = st.columns(4)
    with c_info1:
        st.metric("Budget horaire max", f"${max_hourly:.4f}")
    with c_info2:
        st.metric("Budget mensuel équivalent", f"${budget_monthly:,.2f}")
    with c_info3:
        st.metric("Instances filtrées", f"{len(reco_df):,}")
    with c_info4:
        best_price = top_reco["coût_horaire"].min()
        st.metric("Meilleur prix trouvé", f"${best_price:.4f}/h")

    st.markdown("---")

    for i, row in top_reco.head(3).iterrows():
        flag = REGION_FLAGS.get(row.get("Region Code", ""), "🌐")
        medal = ["🥇", "🥈", "🥉"][i]
        econ = row["économie_vs_od"]
        econ_str = f"−{econ:.1f}% vs On-Demand" if econ > 0 else "Tarif On-Demand"
        with st.container():
            st.markdown(f"### {medal} {row['Ec2Type']} — {flag} {row['Region Name']}")
            col_m1, col_m2, col_m3, col_m4, col_m5 = st.columns(5)
            with col_m1:
                st.metric("Coût horaire", f"${row['coût_horaire']:.4f}")
            with col_m2:
                st.metric("Coût mensuel", f"${row['coût_mensuel']:,.2f}")
            with col_m3:
                st.metric("vCPU", int(row["CPU"]))
            with col_m4:
                st.metric("RAM", f"{row['RAM']} Go")
            with col_m5:
                st.metric("Économie", econ_str)
            saps_val = int(row["saps_norm"]) if row["saps_norm"] > 0 else "N/A"
            st.caption(f"OS : {row['OS']} · SAPS : {saps_val} · Engagement : {engagement}")
            st.markdown("---")

    st.subheader("📊 Comparaison visuelle des recommandations")
    tab_chart1, tab_chart2, tab_chart3 = st.tabs(["💰 Coûts", "🔥 Performance vs Prix", "🌍 Par région"])

    with tab_chart1:
        chart_df = top_reco[["Ec2Type","Region Name","coût_mensuel","coût_annuel","CPU","RAM"]].copy()
        chart_df["Label"] = chart_df["Ec2Type"] + "\n" + chart_df["Region Name"]
        fig_bar_reco = px.bar(
            chart_df,
            x="Label",
            y="coût_mensuel",
            color="Region Name",
            title="Coût mensuel des instances recommandées",
            labels={"coût_mensuel": "Coût mensuel ($)", "Label": "Instance"},
            color_discrete_sequence=px.colors.qualitative.Set2,
            text_auto=".2f"
        )
        fig_bar_reco.update_traces(texttemplate="$%{y:.2f}", textposition="outside")
        fig_bar_reco.update_layout(height=450, xaxis_tickangle=-30, showlegend=True)
        if max_hourly * 730 > 0:
            fig_bar_reco.add_hline(y=budget_monthly, line_dash="dash", line_color="red",
                                   annotation_text=f"Budget: ${budget_monthly:,.0f}", annotation_position="top right")
        st.plotly_chart(fig_bar_reco, width="stretch")

    with tab_chart2:
        scatter_reco = top_reco[top_reco["saps_norm"] > 0].copy()
        if not scatter_reco.empty:
            fig_sc = px.scatter(
                scatter_reco,
                x="coût_horaire",
                y="saps_norm",
                size="RAM",
                color="Region Name",
                text="Ec2Type",
                title="Performance (SAPS) vs Coût horaire",
                labels={"coût_horaire": "Coût $/h", "saps_norm": "SAPS (performance)", "RAM": "RAM (Go)"},
                color_discrete_sequence=px.colors.qualitative.Bold
            )
            fig_sc.update_traces(textposition="top center")
            fig_sc.update_layout(height=500)
            st.plotly_chart(fig_sc, width="stretch")
            st.caption("Les instances en haut à gauche offrent le meilleur rapport performance/prix.")
        else:
            st.info("Données SAPS non disponibles pour les instances sélectionnées.")

    with tab_chart3:
        region_reco = top_reco.groupby("Region Name").agg(
            nb_instances=("Ec2Type", "count"),
            coût_min=("coût_mensuel", "min"),
            coût_moy=("coût_mensuel", "mean"),
            coût_max=("coût_mensuel", "max"),
        ).reset_index()
        fig_reg = px.bar(
            region_reco,
            x="Region Name",
            y=["coût_min", "coût_moy", "coût_max"],
            barmode="group",
            title="Coûts mensuels min / moyen / max par région (recommandations)",
            labels={"value": "Coût mensuel ($)", "Region Name": "Région", "variable": "Indicateur"},
            color_discrete_map={"coût_min": "#00CC96", "coût_moy": "#636EFA", "coût_max": "#EF553B"}
        )
        newnames = {"coût_min": "Minimum", "coût_moy": "Moyen", "coût_max": "Maximum"}
        fig_reg.for_each_trace(lambda t: t.update(name=newnames.get(t.name, t.name)))
        fig_reg.update_layout(height=400, xaxis_tickangle=-30)
        st.plotly_chart(fig_reg, width="stretch")

    st.markdown("---")
    st.subheader("📋 Tableau complet des recommandations")
    table_cols = ["Region Name", "Ec2Type", "OS", "CPU", "RAM", "coût_horaire", "coût_mensuel", "coût_annuel", "économie_vs_od", "saps_norm"]
    available_table_cols = [c for c in table_cols if c in top_reco.columns]
    display_reco = top_reco[available_table_cols].copy()
    display_reco.columns = ["Région", "Type EC2", "OS", "vCPU", "RAM (Go)", "$/h", "$/mois", "$/an", "Économie (%)", "SAPS"][:len(available_table_cols)]
    display_reco = display_reco.reset_index(drop=True)
    display_reco.index = display_reco.index + 1

    def highlight_top3(row):
        if row.name <= 3:
            return ["background-color: #fffbcc"] * len(row)
        return [""] * len(row)

    st.dataframe(
        display_reco.style.apply(highlight_top3, axis=1).format({
            "$/h": "${:.4f}",
            "$/mois": "${:,.2f}",
            "$/an": "${:,.2f}",
            "Économie (%)": "{:.1f}%",
        }, na_rep="N/A"),
        width="stretch"
    )

    st.markdown("---")
    st.subheader("📦 Matrice de décision multi-critères")
    st.markdown("Score composite normalisé (0–100) tenant compte du prix, des ressources et des performances.")

    score_df = top_reco.copy()
    max_saps = score_df["saps_norm"].max() or 1
    max_cpu = score_df["CPU"].max() or 1
    max_ram = score_df["RAM"].max() or 1
    min_cost = score_df["coût_horaire"].min() or 0.0001

    score_df["score_prix"] = (min_cost / score_df["coût_horaire"] * 100).clip(0, 100)
    score_df["score_cpu"] = (score_df["CPU"] / max_cpu * 100).clip(0, 100)
    score_df["score_ram"] = (score_df["RAM"] / max_ram * 100).clip(0, 100)
    score_df["score_perf"] = (score_df["saps_norm"] / max_saps * 100).clip(0, 100)
    score_df["score_global"] = (
        score_df["score_prix"] * 0.35 +
        score_df["score_cpu"] * 0.25 +
        score_df["score_ram"] * 0.20 +
        score_df["score_perf"] * 0.20
    ).round(1)
    score_df["Label"] = score_df["Ec2Type"] + " / " + score_df["Region Name"].str.split("(").str[0].str.strip()

    fig_radar_list = []
    categories = ["Prix", "vCPU", "RAM", "Performance"]
    for _, row_r in score_df.head(5).iterrows():
        fig_radar_list.append({
            "name": row_r["Label"],
            "values": [row_r["score_prix"], row_r["score_cpu"], row_r["score_ram"], row_r["score_perf"]]
        })

    fig_radar = go.Figure()
    colors_radar = px.colors.qualitative.Set2
    for idx, item in enumerate(fig_radar_list):
        vals = item["values"] + [item["values"][0]]
        cats = categories + [categories[0]]
        fig_radar.add_trace(go.Scatterpolar(
            r=vals,
            theta=cats,
            fill="toself",
            name=item["name"],
            line_color=colors_radar[idx % len(colors_radar)],
            opacity=0.7
        ))
    fig_radar.update_layout(
        polar=dict(radialaxis=dict(visible=True, range=[0, 100])),
        title="Radar multi-critères des 5 meilleures instances",
        height=500,
        showlegend=True
    )
    st.plotly_chart(fig_radar, width="stretch")

    score_table = score_df[["Label", "score_prix", "score_cpu", "score_ram", "score_perf", "score_global"]].copy()
    score_table.columns = ["Instance / Région", "Score Prix", "Score CPU", "Score RAM", "Score Perf.", "Score Global"]
    score_table = score_table.sort_values("Score Global", ascending=False).reset_index(drop=True)
    score_table.index = score_table.index + 1

    def color_score(val):
        if isinstance(val, float):
            if val >= 75:
                return "background-color: #d4edda; color: #155724"
            elif val >= 50:
                return "background-color: #fff3cd; color: #856404"
            else:
                return "background-color: #f8d7da; color: #721c24"
        return ""

    st.dataframe(
        score_table.style.applymap(color_score, subset=["Score Prix", "Score CPU", "Score RAM", "Score Perf.", "Score Global"])
                         .format("{:.1f}", subset=["Score Prix", "Score CPU", "Score RAM", "Score Perf.", "Score Global"]),
        width="stretch"
    )
    st.caption("Score Global = 35% Prix + 25% CPU + 20% RAM + 20% Performance. Les 3 premiers sont surlignés en jaune dans le tableau des recommandations.")

    st.markdown("---")
    st.subheader("📥 Exporter les résultats")
    st.markdown("Téléchargez les recommandations dans le format de votre choix.")

    export_df = top_reco[[
        "Region Name", "Ec2Type", "OS", "CPU", "RAM",
        "coût_horaire", "coût_mensuel", "coût_annuel", "économie_vs_od", "saps_norm"
    ]].copy()
    export_df.columns = [
        "Région", "Type EC2", "OS", "vCPU", "RAM (Go)",
        "Coût $/h", "Coût $/mois", "Coût $/an", "Économie vs OD (%)", "SAPS"
    ]
    export_df = export_df.reset_index(drop=True)
    export_df.index = export_df.index + 1

    score_export = score_table.copy()

    def build_excel(reco_data, score_data, criteria):
        wb = openpyxl.Workbook()

        HDR_FILL = PatternFill("solid", fgColor="232F3E")
        HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
        GOLD_FILL = PatternFill("solid", fgColor="FFF3CD")
        GREEN_FILL = PatternFill("solid", fgColor="D4EDDA")
        YELLOW_FILL = PatternFill("solid", fgColor="FFF3CD")
        RED_FILL = PatternFill("solid", fgColor="F8D7DA")
        BORDER = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )
        CENTER = Alignment(horizontal="center", vertical="center")
        LEFT = Alignment(horizontal="left", vertical="center")

        ws1 = wb.active
        ws1.title = "Recommandations"

        ws1.merge_cells("A1:J1")
        title_cell = ws1["A1"]
        title_cell.value = "☁️  Recommandations AWS EC2 — Ressources AWS"
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor="FF9900")
        title_cell.alignment = CENTER
        ws1.row_dimensions[1].height = 28

        ws1.merge_cells("A2:J2")
        sub_cell = ws1["A2"]
        sub_cell.value = (
            f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}  |  "
            f"Budget : {criteria['budget']}  |  Engagement : {criteria['engagement']}  |  "
            f"vCPU min : {criteria['cpu']}  |  RAM min : {criteria['ram']} Go  |  "
            f"Priorité : {criteria['prio']}"
        )
        sub_cell.font = Font(italic=True, size=9, color="555555")
        sub_cell.alignment = LEFT
        ws1.row_dimensions[2].height = 18

        headers = list(reco_data.columns)
        for col_idx, hdr in enumerate(headers, start=1):
            cell = ws1.cell(row=4, column=col_idx, value=hdr)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = CENTER
            cell.border = BORDER
        ws1.row_dimensions[4].height = 20

        money_fmt = '#,##0.0000 "$"'
        money_mrr = '#,##0.00 "$"'
        pct_fmt = '0.0"%"'

        col_formats = {
            "Coût $/h": money_fmt,
            "Coût $/mois": money_mrr,
            "Coût $/an": money_mrr,
            "Économie vs OD (%)": pct_fmt,
            "SAPS": '#,##0',
            "vCPU": '0',
            "RAM (Go)": '0',
        }

        for row_idx, row in reco_data.iterrows():
            excel_row = row_idx + 4
            row_fill = GOLD_FILL if row_idx <= 3 else PatternFill("solid", fgColor="FFFFFF" if row_idx % 2 else "F8F9FA")
            for col_idx, (col_name, val) in enumerate(zip(headers, row), start=1):
                cell = ws1.cell(row=excel_row, column=col_idx, value=val)
                cell.fill = row_fill
                cell.border = BORDER
                cell.alignment = CENTER
                fmt = col_formats.get(col_name)
                if fmt:
                    cell.number_format = fmt
            ws1.row_dimensions[excel_row].height = 16

        col_widths = [28, 18, 10, 8, 10, 14, 14, 14, 16, 10]
        for i, w in enumerate(col_widths, start=1):
            ws1.column_dimensions[get_column_letter(i)].width = w

        ws2 = wb.create_sheet("Scores Multi-Critères")
        ws2.merge_cells("A1:F1")
        t2 = ws2["A1"]
        t2.value = "Matrice de Décision Multi-Critères"
        t2.font = Font(bold=True, size=13, color="FFFFFF")
        t2.fill = PatternFill("solid", fgColor="232F3E")
        t2.alignment = CENTER
        ws2.row_dimensions[1].height = 24

        score_headers = list(score_data.columns)
        for col_idx, hdr in enumerate(score_headers, start=1):
            cell = ws2.cell(row=3, column=col_idx, value=hdr)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = CENTER
            cell.border = BORDER
        ws2.row_dimensions[3].height = 18

        score_cols_num = ["Score Prix", "Score CPU", "Score RAM", "Score Perf.", "Score Global"]
        for row_idx, row in score_data.iterrows():
            excel_row = row_idx + 3
            for col_idx, (col_name, val) in enumerate(zip(score_headers, row), start=1):
                cell = ws2.cell(row=excel_row, column=col_idx, value=val)
                cell.border = BORDER
                cell.alignment = CENTER
                if col_name in score_cols_num and isinstance(val, float):
                    cell.number_format = "0.0"
                    if val >= 75:
                        cell.fill = GREEN_FILL
                        cell.font = Font(color="155724")
                    elif val >= 50:
                        cell.fill = YELLOW_FILL
                        cell.font = Font(color="856404")
                    else:
                        cell.fill = RED_FILL
                        cell.font = Font(color="721C24")
            ws2.row_dimensions[excel_row].height = 16

        ws2.column_dimensions["A"].width = 40
        for i in range(2, 7):
            ws2.column_dimensions[get_column_letter(i)].width = 14

        ws3 = wb.create_sheet("Glossaire")
        ws3.merge_cells("A1:B1")
        g1 = ws3["A1"]
        g1.value = "Glossaire des indicateurs"
        g1.font = Font(bold=True, size=12, color="FFFFFF")
        g1.fill = PatternFill("solid", fgColor="FF9900")
        g1.alignment = CENTER
        ws3.row_dimensions[1].height = 22

        glossaire = [
            ("Coût $/h", "Coût horaire de l'instance selon le type d'engagement choisi"),
            ("Coût $/mois", "Coût mensuel estimé sur 730 heures"),
            ("Coût $/an", "Coût annuel estimé sur 8 760 heures"),
            ("Économie vs OD (%)", "Pourcentage d'économie par rapport au tarif On-Demand"),
            ("SAPS", "Standard Application Performance Standard — mesure de performance SAP"),
            ("Score Prix", "Score normalisé 0–100 : 100 = moins cher de la sélection"),
            ("Score CPU", "Score normalisé 0–100 : 100 = plus de vCPU de la sélection"),
            ("Score RAM", "Score normalisé 0–100 : 100 = plus de RAM de la sélection"),
            ("Score Perf.", "Score normalisé 0–100 basé sur les SAPS"),
            ("Score Global", "35% Prix + 25% CPU + 20% RAM + 20% Performance"),
        ]
        for r, (terme, definition) in enumerate(glossaire, start=3):
            c1 = ws3.cell(row=r, column=1, value=terme)
            c1.font = Font(bold=True)
            c1.border = BORDER
            c1.fill = PatternFill("solid", fgColor="F2F2F2" if r % 2 else "FFFFFF")
            c2 = ws3.cell(row=r, column=2, value=definition)
            c2.border = BORDER
            c2.fill = PatternFill("solid", fgColor="F2F2F2" if r % 2 else "FFFFFF")
            ws3.row_dimensions[r].height = 15
        ws3.column_dimensions["A"].width = 22
        ws3.column_dimensions["B"].width = 65

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    class AWSPDF(FPDF):
        def header(self):
            self.set_fill_color(35, 47, 62)
            self.rect(0, 0, 210, 18, "F")
            self.set_font("Helvetica", "B", 13)
            self.set_text_color(255, 153, 0)
            self.set_xy(0, 3)
            self.cell(0, 12, "  Ressources AWS — Recommandations EC2", align="L")
            self.set_font("Helvetica", "", 8)
            self.set_text_color(180, 180, 180)
            self.set_xy(0, 3)
            self.cell(0, 12, datetime.now().strftime("%d/%m/%Y  "), align="R")
            self.ln(14)

        def footer(self):
            self.set_y(-12)
            self.set_font("Helvetica", "I", 8)
            self.set_text_color(150, 150, 150)
            self.set_fill_color(245, 245, 245)
            self.rect(0, self.get_y(), 210, 12, "F")
            self.cell(0, 10, f"Page {self.page_no()} — Données AWS Pricing Avril 2026", align="C")

    def build_pdf(reco_data, score_data, criteria):
        pdf = AWSPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        pdf.set_font("Helvetica", "B", 16)
        pdf.set_text_color(35, 47, 62)
        pdf.cell(0, 10, "Rapport de Recommandations AWS EC2", ln=True, align="C")
        pdf.ln(2)

        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(80, 80, 80)
        pdf.set_fill_color(255, 249, 230)
        pdf.set_draw_color(255, 153, 0)
        pdf.rect(10, pdf.get_y(), 190, 22, "FD")
        pdf.set_xy(12, pdf.get_y() + 2)
        pdf.multi_cell(186, 5,
            f"Budget : {criteria['budget']}   |   Engagement : {criteria['engagement']}   |   "
            f"vCPU min : {criteria['cpu']}   |   RAM min : {criteria['ram']} Go\n"
            f"Priorité : {criteria['prio']}   |   Instances analysées : {criteria['total_filtered']:,}   |   "
            f"Résultats affichés : {len(reco_data)}",
            align="L"
        )
        pdf.ln(6)

        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(35, 47, 62)
        pdf.cell(0, 8, "Meilleures instances recommandées", ln=True)
        pdf.ln(1)

        col_widths = [40, 25, 18, 12, 14, 20, 22, 22]
        col_names = ["Région", "Type EC2", "OS", "vCPU", "RAM (Go)", "$/h", "$/mois", "$/an"]

        pdf.set_fill_color(35, 47, 62)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 8)
        for w, name in zip(col_widths, col_names):
            pdf.cell(w, 7, name, border=1, align="C", fill=True)
        pdf.ln()

        medals = {1: "1.", 2: "2.", 3: "3."}
        for i, row in reco_data.iterrows():
            pdf.set_font("Helvetica", "B" if i <= 3 else "", 7.5)
            if i <= 3:
                pdf.set_fill_color(255, 249, 230)
                pdf.set_text_color(101, 68, 0)
            else:
                bg = 248 if i % 2 == 0 else 255
                pdf.set_fill_color(bg, bg, bg)
                pdf.set_text_color(40, 40, 40)
            medal = medals.get(i, "")
            vals = [
                f"{medal} {row['Région']}"[:22],
                str(row["Type EC2"]),
                str(row["OS"]),
                str(int(row["vCPU"])),
                f"{row['RAM (Go)']} Go",
                f"${row['Coût $/h']:.4f}",
                f"${row['Coût $/mois']:,.2f}",
                f"${row['Coût $/an']:,.2f}",
            ]
            for w, val in zip(col_widths, vals):
                pdf.cell(w, 6, val, border=1, align="C", fill=True)
            pdf.ln()

        pdf.ln(8)
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(35, 47, 62)
        pdf.cell(0, 8, "Matrice de Décision Multi-Critères", ln=True)
        pdf.ln(1)

        score_col_widths = [70, 22, 22, 22, 22, 22]
        score_col_names = ["Instance / Région", "Prix", "CPU", "RAM", "Perf.", "Global"]

        pdf.set_fill_color(35, 47, 62)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 8)
        for w, name in zip(score_col_widths, score_col_names):
            pdf.cell(w, 7, name, border=1, align="C", fill=True)
        pdf.ln()

        for _, row in score_data.iterrows():
            glob = row.get("Score Global", 0)
            if glob >= 75:
                pdf.set_fill_color(212, 237, 218)
                pdf.set_text_color(21, 87, 36)
            elif glob >= 50:
                pdf.set_fill_color(255, 243, 205)
                pdf.set_text_color(133, 100, 4)
            else:
                pdf.set_fill_color(248, 215, 218)
                pdf.set_text_color(114, 28, 36)
            pdf.set_font("Helvetica", "", 7.5)
            scores = [
                str(row.get("Instance / Région", ""))[:35],
                f"{row.get('Score Prix', 0):.1f}",
                f"{row.get('Score CPU', 0):.1f}",
                f"{row.get('Score RAM', 0):.1f}",
                f"{row.get('Score Perf.', 0):.1f}",
                f"{glob:.1f}",
            ]
            for w, val in zip(score_col_widths, scores):
                pdf.cell(w, 6, val, border=1, align="C", fill=True)
            pdf.ln()

        pdf.ln(8)
        pdf.set_font("Helvetica", "I", 8)
        pdf.set_text_color(100, 100, 100)
        pdf.multi_cell(0, 5,
            "Score Global = 35% Prix + 25% vCPU + 20% RAM + 20% Performance (SAPS). "
            "Scores normalisés sur 100 par rapport à la sélection affichée. "
            "Les 3 premières lignes du tableau des recommandations correspondent aux meilleures instances selon vos critères.",
            align="L"
        )

        return bytes(pdf.output())

    criteria_info = {
        "budget": f"{budget_val:,.2f} {budget_type}",
        "engagement": engagement,
        "cpu": min_cpu,
        "ram": min_ram,
        "prio": prioritize,
        "total_filtered": len(reco_df),
    }

    col_ex1, col_ex2, col_ex3 = st.columns([1, 1, 2])

    with col_ex1:
        with st.spinner("Préparation de l'Excel..."):
            excel_bytes = build_excel(export_df, score_table, criteria_info)
        fname_xlsx = f"recommandations_aws_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.download_button(
            label="📊 Télécharger Excel (.xlsx)",
            data=excel_bytes,
            file_name=fname_xlsx,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.caption("3 onglets : Recommandations · Scores · Glossaire")

    with col_ex2:
        with st.spinner("Préparation du PDF..."):
            pdf_bytes = build_pdf(export_df, score_table, criteria_info)
        fname_pdf = f"recommandations_aws_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        st.download_button(
            label="📄 Télécharger PDF (.pdf)",
            data=pdf_bytes,
            file_name=fname_pdf,
            mime="application/pdf",
            use_container_width=True,
        )
        st.caption("Rapport formaté prêt à partager")

    with col_ex3:
        st.info(
            "💡 **Astuce** : modifiez vos critères en haut de page, "
            "les exports se mettent à jour automatiquement avec les nouvelles recommandations."
        )


elif page == "⚖️ Comparaison TCO":
    st.title("⚖️ Comparaison & TCO")
    st.markdown("Comparez deux instances EC2 côte à côte et estimez leur Coût Total de Possession (TCO) sur 1, 2 et 3 ans.")
    st.markdown("---")

    all_types = sorted(df["Ec2Type"].dropna().unique())
    all_regions = sorted(df["Region Name"].dropna().unique())

    st.subheader("🔍 Sélectionnez les deux instances à comparer")
    col_l, col_sep, col_r = st.columns([5, 1, 5])

    with col_l:
        st.markdown("#### Instance A")
        reg_a = st.selectbox("Région A", all_regions, index=0, key="tco_reg_a")
        os_opts_a = sorted(df[df["Region Name"] == reg_a]["OS"].dropna().unique())
        os_a = st.selectbox("OS A", os_opts_a, key="tco_os_a")
        types_a = sorted(df[(df["Region Name"] == reg_a) & (df["OS"] == os_a)]["Ec2Type"].dropna().unique())
        type_a = st.selectbox("Type EC2 A", types_a, key="tco_type_a")

    with col_sep:
        st.markdown("<div style='text-align:center; font-size:2.5rem; margin-top:4rem;'>⚡</div>", unsafe_allow_html=True)

    with col_r:
        st.markdown("#### Instance B")
        reg_b = st.selectbox("Région B", all_regions, index=min(1, len(all_regions)-1), key="tco_reg_b")
        os_opts_b = sorted(df[df["Region Name"] == reg_b]["OS"].dropna().unique())
        os_b = st.selectbox("OS B", os_opts_b, key="tco_os_b")
        types_b = sorted(df[(df["Region Name"] == reg_b) & (df["OS"] == os_b)]["Ec2Type"].dropna().unique())
        type_b = st.selectbox("Type EC2 B", types_b, key="tco_type_b")

    st.markdown("---")
    st.subheader("⚙️ Paramètres TCO")
    col_p1, col_p2, col_p3 = st.columns(3)
    with col_p1:
        nb_instances = st.number_input("Nombre d'instances", min_value=1, max_value=10000, value=5, key="tco_nb")
        usage_pct = st.slider("Taux d'utilisation (%)", 10, 100, 100, 5, key="tco_usage",
                              help="100% = instance allumée 24h/24. 50% = allumée 12h/jour en moyenne.")
    with col_p2:
        ebs_vol_gb = st.number_input("Volume EBS gp3 par instance (Go)", min_value=0, value=100, step=50, key="tco_ebs")
        s3_vol_tb = st.number_input("Volume S3 total (To)", min_value=0.0, value=5.0, step=1.0, key="tco_s3")
    with col_p3:
        engagement = st.selectbox(
            "Engagement tarifaire",
            ["On-Demand (PayAsYouGo)", "Reserved 1 an", "Reserved 3 ans", "Savings Plan (~1 an)"],
            key="tco_engagement"
        )
        transfer_gb = st.number_input("Transfert sortant mensuel (Go)", min_value=0, value=100, step=50, key="tco_transfer",
                                      help="Données transférées hors AWS (~0.09$/Go en Europe)")

    cost_col_map = {
        "On-Demand (PayAsYouGo)": "ODH",
        "Reserved 1 an": "CostH1yr",
        "Reserved 3 ans": "CostH3yr",
        "Savings Plan (~1 an)": "SavingsPlanH",
    }
    cost_col = cost_col_map[engagement]

    row_a = df[(df["Region Name"] == reg_a) & (df["OS"] == os_a) & (df["Ec2Type"] == type_a)]
    row_b = df[(df["Region Name"] == reg_b) & (df["OS"] == os_b) & (df["Ec2Type"] == type_b)]

    if row_a.empty or row_b.empty:
        st.warning("Instance introuvable. Vérifiez vos sélections.")
        st.stop()

    ra = row_a.iloc[0]
    rb = row_b.iloc[0]

    TRANSFER_COST = 0.09
    hours_per_year = 8760 * (usage_pct / 100)
    hours_per_month = 730 * (usage_pct / 100)

    def tco_for(row, years):
        ec2_h = (row.get(cost_col) or 0) * hours_per_year * years * nb_instances
        ebs_gp3 = (row.get("gp3G/M") or 0) * ebs_vol_gb * nb_instances * 12 * years
        s3_gb = s3_vol_tb * 1024
        s3_cost_m = s3_gb * (row.get("s3<50T") or 0) if s3_gb <= 50*1024 else (
            50*1024*(row.get("s3<50T") or 0) + (s3_gb - 50*1024)*(row.get("s3<450T") or 0)
        )
        s3_total = s3_cost_m * 12 * years
        transfer_total = transfer_gb * TRANSFER_COST * 12 * years
        total = ec2_h + ebs_gp3 + s3_total + transfer_total
        return {
            "ec2": ec2_h,
            "ebs": ebs_gp3,
            "s3": s3_total,
            "transfer": transfer_total,
            "total": total,
        }

    tco_a1 = tco_for(ra, 1); tco_a2 = tco_for(ra, 2); tco_a3 = tco_for(ra, 3)
    tco_b1 = tco_for(rb, 1); tco_b2 = tco_for(rb, 2); tco_b3 = tco_for(rb, 3)

    # ── Fiches récapitulatives ─────────────────────────────────────────────
    st.markdown("---")
    st.subheader("📋 Fiches récapitulatives")
    card_a, card_sep2, card_b = st.columns([5, 1, 5])

    def render_card(row, label, tco1, tco2, tco3, cost_col):
        flag = REGION_FLAGS.get(row.get("Region Code", ""), "🌐")
        od = row.get("ODH") or 0
        cph = row.get(cost_col) or 0
        saving = (1 - cph / od) * 100 if od > 0 else 0
        st.markdown(f"### {label}")
        st.markdown(f"**{flag} {row['Region Name']}** · `{row['Ec2Type']}` · {row['OS']}")
        m1, m2, m3 = st.columns(3)
        m1.metric("vCPU", int(row["CPU"]) if pd.notna(row["CPU"]) else "—")
        m2.metric("RAM", f"{row['RAM']} Go" if pd.notna(row["RAM"]) else "—")
        m3.metric("SAPS", f"{int(row['SAPS']):,}" if pd.notna(row.get("SAPS")) and row.get("SAPS", 0) > 0 else "N/A")
        n1, n2, n3 = st.columns(3)
        n1.metric("$/h", f"${cph:.4f}")
        n2.metric("$/mois", f"${cph*730:,.2f}")
        n3.metric("Économie vs OD", f"{saving:.1f}%")
        st.markdown(f"**TCO ({nb_instances} inst. · {usage_pct}% utilisation)**")
        t1, t2, t3 = st.columns(3)
        t1.metric("1 an", f"${tco1['total']:,.0f}")
        t2.metric("2 ans", f"${tco2['total']:,.0f}")
        t3.metric("3 ans", f"${tco3['total']:,.0f}")

    with card_a:
        render_card(ra, "🅰 Instance A", tco_a1, tco_a2, tco_a3, cost_col)
    with card_sep2:
        st.markdown("")
    with card_b:
        render_card(rb, "🅱 Instance B", tco_b1, tco_b2, tco_b3, cost_col)

    # ── Gagnant ────────────────────────────────────────────────────────────
    st.markdown("---")
    winner_1 = "A" if tco_a1["total"] <= tco_b1["total"] else "B"
    winner_3 = "A" if tco_a3["total"] <= tco_b3["total"] else "B"
    diff_1 = abs(tco_a1["total"] - tco_b1["total"])
    diff_3 = abs(tco_a3["total"] - tco_b3["total"])
    pct_1 = diff_1 / max(tco_a1["total"], tco_b1["total"]) * 100
    pct_3 = diff_3 / max(tco_a3["total"], tco_b3["total"]) * 100

    col_w1, col_w2 = st.columns(2)
    with col_w1:
        icon = "🅰" if winner_1 == "A" else "🅱"
        st.success(f"**Gagnant à 1 an : Instance {winner_1} {icon}**  \n"
                   f"Économie de **${diff_1:,.0f}** ({pct_1:.1f}% moins cher)")
    with col_w2:
        icon = "🅰" if winner_3 == "A" else "🅱"
        st.success(f"**Gagnant à 3 ans : Instance {winner_3} {icon}**  \n"
                   f"Économie de **${diff_3:,.0f}** ({pct_3:.1f}% moins cher)")

    # ── Graphiques ─────────────────────────────────────────────────────────
    st.markdown("---")
    st.subheader("📊 Visualisations comparatives")
    tab1, tab2, tab3, tab4 = st.tabs(["📈 TCO cumulé", "🧩 Répartition des coûts", "📊 Comparaison specs", "💹 Seuil de rentabilité"])

    with tab1:
        months = list(range(1, 37))
        tco_a_monthly = [tco_for(ra, m/12)["total"] for m in months]
        tco_b_monthly = [tco_for(rb, m/12)["total"] for m in months]

        fig_cum = go.Figure()
        fig_cum.add_trace(go.Scatter(
            x=months, y=tco_a_monthly,
            mode="lines+markers", name=f"Instance A — {type_a}",
            line=dict(color="#636EFA", width=2.5),
            marker=dict(size=4),
        ))
        fig_cum.add_trace(go.Scatter(
            x=months, y=tco_b_monthly,
            mode="lines+markers", name=f"Instance B — {type_b}",
            line=dict(color="#EF553B", width=2.5),
            marker=dict(size=4),
        ))
        for yr_m, label in [(12, "1 an"), (24, "2 ans"), (36, "3 ans")]:
            fig_cum.add_vline(x=yr_m, line_dash="dot", line_color="gray",
                              annotation_text=label, annotation_position="top right")
        fig_cum.update_layout(
            title=f"TCO cumulé sur 36 mois — {nb_instances} instance(s) · {usage_pct}% utilisation",
            xaxis_title="Mois", yaxis_title="Coût cumulé ($)",
            height=450, hovermode="x unified",
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
        )
        st.plotly_chart(fig_cum, width="stretch")

    with tab2:
        composants = ["EC2", "EBS", "S3", "Transfert"]
        vals_a3 = [tco_a3["ec2"], tco_a3["ebs"], tco_a3["s3"], tco_a3["transfer"]]
        vals_b3 = [tco_b3["ec2"], tco_b3["ebs"], tco_b3["s3"], tco_b3["transfer"]]

        fig_pie2 = make_subplots(rows=1, cols=2, specs=[[{"type":"pie"}, {"type":"pie"}]],
                                 subplot_titles=[f"Instance A — {type_a}", f"Instance B — {type_b}"])
        palette = px.colors.qualitative.Set2
        non_zero_a = [(c, v) for c, v in zip(composants, vals_a3) if v > 0]
        non_zero_b = [(c, v) for c, v in zip(composants, vals_b3) if v > 0]

        fig_pie2.add_trace(go.Pie(
            labels=[c for c, _ in non_zero_a], values=[v for _, v in non_zero_a],
            hole=0.42, marker_colors=palette,
            textinfo="label+percent", hovertemplate="<b>%{label}</b><br>$%{value:,.0f}<extra></extra>"
        ), row=1, col=1)
        fig_pie2.add_trace(go.Pie(
            labels=[c for c, _ in non_zero_b], values=[v for _, v in non_zero_b],
            hole=0.42, marker_colors=palette,
            textinfo="label+percent", hovertemplate="<b>%{label}</b><br>$%{value:,.0f}<extra></extra>"
        ), row=1, col=2)
        fig_pie2.update_layout(title="Répartition du TCO sur 3 ans par composant", height=400,
                               legend=dict(orientation="h"))
        st.plotly_chart(fig_pie2, width="stretch")

        bar_data = pd.DataFrame({
            "Composant": composants * 2,
            "Instance": [f"A — {type_a}"] * 4 + [f"B — {type_b}"] * 4,
            "Coût ($)": vals_a3 + vals_b3,
        })
        fig_bar_comp = px.bar(bar_data, x="Composant", y="Coût ($)", color="Instance",
                              barmode="group", title="Détail des coûts TCO 3 ans par composant",
                              color_discrete_sequence=["#636EFA", "#EF553B"])
        fig_bar_comp.update_layout(height=380)
        st.plotly_chart(fig_bar_comp, width="stretch")

    with tab3:
        specs_labels = ["vCPU", "RAM (Go)", "SAPS"]
        specs_a = [
            float(ra["CPU"]) if pd.notna(ra.get("CPU")) else 0,
            float(ra["RAM"]) if pd.notna(ra.get("RAM")) else 0,
            float(ra["SAPS"]) if pd.notna(ra.get("SAPS")) and ra.get("SAPS", 0) > 0 else 0,
        ]
        specs_b = [
            float(rb["CPU"]) if pd.notna(rb.get("CPU")) else 0,
            float(rb["RAM"]) if pd.notna(rb.get("RAM")) else 0,
            float(rb["SAPS"]) if pd.notna(rb.get("SAPS")) and rb.get("SAPS", 0) > 0 else 0,
        ]
        max_vals = [max(a, b) or 1 for a, b in zip(specs_a, specs_b)]
        norm_a = [v / m * 100 for v, m in zip(specs_a, max_vals)]
        norm_b = [v / m * 100 for v, m in zip(specs_b, max_vals)]

        fig_specs = go.Figure()
        fig_specs.add_trace(go.Bar(
            name=f"Instance A — {type_a}", x=specs_labels, y=specs_a,
            marker_color="#636EFA",
            text=[f"{v:,.0f}" for v in specs_a], textposition="outside"
        ))
        fig_specs.add_trace(go.Bar(
            name=f"Instance B — {type_b}", x=specs_labels, y=specs_b,
            marker_color="#EF553B",
            text=[f"{v:,.0f}" for v in specs_b], textposition="outside"
        ))
        fig_specs.update_layout(barmode="group", title="Comparaison des spécifications techniques",
                                height=420, legend=dict(orientation="h"))
        st.plotly_chart(fig_specs, width="stretch")

        cat_radar = ["vCPU", "RAM", "SAPS", "Coût/h inv."]
        cph_a = ra.get(cost_col) or 0.0001
        cph_b = rb.get(cost_col) or 0.0001
        max_cost = max(cph_a, cph_b)
        radar_a = norm_a + [(1 - cph_a/max_cost)*100]
        radar_b = norm_b + [(1 - cph_b/max_cost)*100]

        fig_rad = go.Figure()
        for vals, name, color in [
            (radar_a, f"A — {type_a}", "#636EFA"),
            (radar_b, f"B — {type_b}", "#EF553B"),
        ]:
            closed = vals + [vals[0]]
            cats = cat_radar + [cat_radar[0]]
            fig_rad.add_trace(go.Scatterpolar(
                r=closed, theta=cats, fill="toself", name=name,
                line_color=color, opacity=0.7
            ))
        fig_rad.update_layout(
            polar=dict(radialaxis=dict(visible=True, range=[0, 100])),
            title="Radar specs normalisées (100 = meilleur de la paire)",
            height=450, showlegend=True
        )
        st.plotly_chart(fig_rad, width="stretch")

    with tab4:
        st.markdown("Visualisez à partir de quel mois une instance devient moins coûteuse que l'autre.")
        diff_monthly = [tco_a_monthly[i] - tco_b_monthly[i] for i in range(len(months))]

        crossover = None
        for i in range(1, len(diff_monthly)):
            if diff_monthly[i-1] * diff_monthly[i] <= 0 and diff_monthly[i-1] != diff_monthly[i]:
                crossover = months[i]
                break

        fig_diff = go.Figure()
        colors_diff = ["#636EFA" if d <= 0 else "#EF553B" for d in diff_monthly]
        fig_diff.add_trace(go.Bar(
            x=months, y=diff_monthly,
            marker_color=colors_diff,
            name="Différence TCO (A − B)",
            hovertemplate="Mois %{x}<br>Diff: $%{y:,.0f}<extra></extra>"
        ))
        fig_diff.add_hline(y=0, line_color="black", line_width=1.5)
        if crossover:
            fig_diff.add_vline(x=crossover, line_dash="dash", line_color="green",
                               annotation_text=f"Croisement mois {crossover}",
                               annotation_position="top left")
        fig_diff.update_layout(
            title="Différence de TCO cumulé mois par mois (A − B) — bleu = A moins cher, rouge = B moins cher",
            xaxis_title="Mois", yaxis_title="Différence ($)",
            height=430, showlegend=False
        )
        for yr_m, label in [(12, "1 an"), (24, "2 ans"), (36, "3 ans")]:
            fig_diff.add_vline(x=yr_m, line_dash="dot", line_color="lightgray",
                               annotation_text=label, annotation_position="top right")
        st.plotly_chart(fig_diff, width="stretch")

        if crossover:
            winner_before = "B" if diff_monthly[crossover-1] > 0 else "A"
            winner_after = "A" if winner_before == "B" else "B"
            st.info(f"🔀 **Point de croisement au mois {crossover}** : Instance **{winner_before}** est moins chère avant, "
                    f"Instance **{winner_after}** devient moins chère après.")
        else:
            if diff_monthly[-1] < 0:
                st.info("✅ **Instance A** reste moins chère sur toute la période de 3 ans.")
            else:
                st.info("✅ **Instance B** reste moins chère sur toute la période de 3 ans.")

    # ── Tableau TCO complet ────────────────────────────────────────────────
    st.markdown("---")
    st.subheader("📋 Tableau TCO détaillé")

    tco_table = pd.DataFrame({
        "Période": ["1 an", "2 ans", "3 ans"],
        "A — EC2 ($)": [f"${tco_a1['ec2']:,.0f}", f"${tco_a2['ec2']:,.0f}", f"${tco_a3['ec2']:,.0f}"],
        "A — EBS ($)": [f"${tco_a1['ebs']:,.0f}", f"${tco_a2['ebs']:,.0f}", f"${tco_a3['ebs']:,.0f}"],
        "A — S3 ($)": [f"${tco_a1['s3']:,.0f}", f"${tco_a2['s3']:,.0f}", f"${tco_a3['s3']:,.0f}"],
        "A — Transfert ($)": [f"${tco_a1['transfer']:,.0f}", f"${tco_a2['transfer']:,.0f}", f"${tco_a3['transfer']:,.0f}"],
        "A — TOTAL ($)": [f"${tco_a1['total']:,.0f}", f"${tco_a2['total']:,.0f}", f"${tco_a3['total']:,.0f}"],
        "B — EC2 ($)": [f"${tco_b1['ec2']:,.0f}", f"${tco_b2['ec2']:,.0f}", f"${tco_b3['ec2']:,.0f}"],
        "B — EBS ($)": [f"${tco_b1['ebs']:,.0f}", f"${tco_b2['ebs']:,.0f}", f"${tco_b3['ebs']:,.0f}"],
        "B — S3 ($)": [f"${tco_b1['s3']:,.0f}", f"${tco_b2['s3']:,.0f}", f"${tco_b3['s3']:,.0f}"],
        "B — Transfert ($)": [f"${tco_b1['transfer']:,.0f}", f"${tco_b2['transfer']:,.0f}", f"${tco_b3['transfer']:,.0f}"],
        "B — TOTAL ($)": [f"${tco_b1['total']:,.0f}", f"${tco_b2['total']:,.0f}", f"${tco_b3['total']:,.0f}"],
        "Δ Économie ($)": [
            f"${abs(tco_a1['total']-tco_b1['total']):,.0f} {'→A' if tco_a1['total']<tco_b1['total'] else '→B'}",
            f"${abs(tco_a2['total']-tco_b2['total']):,.0f} {'→A' if tco_a2['total']<tco_b2['total'] else '→B'}",
            f"${abs(tco_a3['total']-tco_b3['total']):,.0f} {'→A' if tco_a3['total']<tco_b3['total'] else '→B'}",
        ],
    })
    st.dataframe(tco_table, width="stretch", hide_index=True)

    # ── Export Excel TCO ───────────────────────────────────────────────────
    st.markdown("---")
    st.subheader("📥 Exporter la comparaison TCO")

    def build_tco_excel(ra, rb, tcos_a, tcos_b, params):
        wb = openpyxl.Workbook()
        HDR_FILL = PatternFill("solid", fgColor="232F3E")
        HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
        ORANGE = PatternFill("solid", fgColor="FF9900")
        BLUE_F = PatternFill("solid", fgColor="DCE6F1")
        RED_F = PatternFill("solid", fgColor="FCE4D6")
        BORDER = Border(
            left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"),
        )
        CENTER = Alignment(horizontal="center", vertical="center")

        ws = wb.active
        ws.title = "TCO Comparaison"

        ws.merge_cells("A1:M1")
        h = ws["A1"]
        h.value = f"⚖️  Comparaison TCO — {ra['Ec2Type']} vs {rb['Ec2Type']}"
        h.font = Font(bold=True, size=13, color="FFFFFF")
        h.fill = ORANGE
        h.alignment = CENTER
        ws.row_dimensions[1].height = 26

        ws.merge_cells("A2:M2")
        s = ws["A2"]
        s.value = (f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}  |  "
                   f"{params['nb']} instance(s)  |  {params['usage']}% utilisation  |  "
                   f"Engagement : {params['engagement']}  |  "
                   f"EBS : {params['ebs']} Go/inst.  |  S3 : {params['s3']} To  |  Transfert : {params['transfer']} Go/mois")
        s.font = Font(italic=True, size=8, color="555555")
        s.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 16

        headers = ["Période", "A — EC2", "A — EBS", "A — S3", "A — Transfert", "A — TOTAL",
                   "B — EC2", "B — EBS", "B — S3", "B — Transfert", "B — TOTAL", "Δ ($)", "Gagnant"]
        for ci, hdr in enumerate(headers, 1):
            c = ws.cell(row=4, column=ci, value=hdr)
            c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = BORDER
        ws.row_dimensions[4].height = 18

        periods = [("1 an", tcos_a[0], tcos_b[0]), ("2 ans", tcos_a[1], tcos_b[1]), ("3 ans", tcos_a[2], tcos_b[2])]
        for ri, (label, ta, tb) in enumerate(periods, 5):
            winner = "A" if ta["total"] <= tb["total"] else "B"
            row_vals = [
                label, ta["ec2"], ta["ebs"], ta["s3"], ta["transfer"], ta["total"],
                tb["ec2"], tb["ebs"], tb["s3"], tb["transfer"], tb["total"],
                abs(ta["total"] - tb["total"]), f"Instance {winner}"
            ]
            for ci, val in enumerate(row_vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = BORDER; c.alignment = CENTER
                if ci in (1,): c.font = Font(bold=True)
                if ci in range(2, 7): c.fill = BLUE_F
                if ci in range(7, 12): c.fill = RED_F
                if isinstance(val, float): c.number_format = '#,##0.00 "$"'
            ws.row_dimensions[ri].height = 15

        col_ws = [10, 14, 14, 14, 16, 16, 14, 14, 14, 16, 16, 14, 14]
        for i, w in enumerate(col_ws, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws2 = wb.create_sheet("Évolution mensuelle")
        ws2.merge_cells("A1:C1")
        h2 = ws2["A1"]
        h2.value = "Évolution TCO mensuelle (36 mois)"
        h2.font = Font(bold=True, size=12, color="FFFFFF")
        h2.fill = HDR_FILL; h2.alignment = CENTER
        ws2.row_dimensions[1].height = 22
        for ci, hdr in enumerate(["Mois", f"TCO A — {ra['Ec2Type']} ($)", f"TCO B — {rb['Ec2Type']} ($)"], 1):
            c = ws2.cell(row=3, column=ci, value=hdr)
            c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = BORDER
        ws2.row_dimensions[3].height = 16
        for m in range(1, 37):
            ta_m = tco_for(ra, m/12)["total"]
            tb_m = tco_for(rb, m/12)["total"]
            for ci, val in enumerate([m, ta_m, tb_m], 1):
                c = ws2.cell(row=m+3, column=ci, value=val)
                c.border = BORDER; c.alignment = CENTER
                if ci in (2, 3): c.number_format = '#,##0.00 "$"'
                if ci == 2: c.fill = BLUE_F
                if ci == 3: c.fill = RED_F
            ws2.row_dimensions[m+3].height = 14
        for i, w in enumerate([8, 28, 28], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    tco_excel = build_tco_excel(
        ra, rb,
        [tco_a1, tco_a2, tco_a3],
        [tco_b1, tco_b2, tco_b3],
        {"nb": nb_instances, "usage": usage_pct, "engagement": engagement,
         "ebs": ebs_vol_gb, "s3": s3_vol_tb, "transfer": transfer_gb}
    )
    fname_tco = f"tco_comparaison_{type_a}_vs_{type_b}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    col_dl1, col_dl2 = st.columns([1, 3])
    with col_dl1:
        st.download_button(
            label="📊 Télécharger TCO Excel",
            data=tco_excel,
            file_name=fname_tco,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.caption("2 onglets : TCO par période · Évolution mensuelle 36 mois")
    with col_dl2:
        st.info(f"💡 Modifiez les sélections d'instances ou les paramètres TCO en haut de page — "
                f"tous les graphiques et l'export se recalculent automatiquement.")


elif page == "🧮 Panier d'instances":
    st.title("🧮 Panier d'instances EC2")
    st.markdown("Constituez votre liste d'instances, visualisez instantanément les coûts cumulés par type et par région.")
    st.markdown("---")

    # ── État du panier dans la session ────────────────────────────────────
    if "panier" not in st.session_state:
        st.session_state.panier = []

    # ── Formulaire d'ajout ────────────────────────────────────────────────
    st.subheader("➕ Ajouter une ligne au panier")
    with st.form("form_panier", clear_on_submit=True):
        fc1, fc2, fc3, fc4, fc5 = st.columns([3, 2, 3, 2, 1])
        with fc1:
            form_region = st.selectbox("Région", sorted(df["Region Name"].unique()), key="f_reg")
        with fc2:
            os_opts = sorted(df[df["Region Name"] == st.session_state.get("f_reg",
                             sorted(df["Region Name"].unique())[0])]["OS"].dropna().unique())
            form_os = st.selectbox("OS", sorted(df["OS"].dropna().unique()), key="f_os")
        with fc3:
            types_dispo = sorted(df["Ec2Type"].dropna().unique())
            form_type = st.selectbox("Type EC2", types_dispo, key="f_type")
        with fc4:
            form_nb = st.number_input("Quantité", min_value=1, max_value=10000, value=1, step=1, key="f_nb")
        with fc5:
            form_engagement = st.selectbox("Engagement", ["OD", "1an", "3ans", "SP"], key="f_eng")
        submitted = st.form_submit_button("➕ Ajouter", use_container_width=True)

    if submitted:
        eng_col = {"OD": "ODH", "1an": "CostH1yr", "3ans": "CostH3yr", "SP": "SavingsPlanH"}[form_engagement]
        match = df[(df["Region Name"] == form_region) & (df["OS"] == form_os) & (df["Ec2Type"] == form_type)]
        if not match.empty:
            row = match.iloc[0]
            cph = float(row.get(eng_col) or row.get("ODH") or 0)
            st.session_state.panier.append({
                "Région": form_region,
                "OS": form_os,
                "Type EC2": form_type,
                "Quantité": int(form_nb),
                "Engagement": form_engagement,
                "$/h unit.": cph,
                "$/h total": cph * form_nb,
                "$/mois total": cph * form_nb * 730,
                "$/an total": cph * form_nb * 8760,
                "vCPU unit.": int(row.get("CPU") or 0),
                "RAM unit. (Go)": float(row.get("RAM") or 0),
            })
        else:
            st.warning("Combinaison région / OS / type introuvable dans les données.")

    # ── Tableau du panier ─────────────────────────────────────────────────
    if st.session_state.panier:
        panier_df = pd.DataFrame(st.session_state.panier)

        # boutons de gestion
        g1, g2, g3 = st.columns([1, 1, 4])
        with g1:
            if st.button("🗑️ Vider le panier", use_container_width=True):
                st.session_state.panier = []
                st.rerun()
        with g2:
            if st.button("🗑️ Supprimer dernière ligne", use_container_width=True):
                st.session_state.panier.pop()
                st.rerun()

        st.markdown("---")

        # ── Histogrammes ──────────────────────────────────────────────────
        st.subheader("📊 Coûts cumulés du panier")

        tab_h1, tab_h2, tab_h3 = st.tabs(["Par type EC2", "Par région", "Par engagement"])

        with tab_h1:
            agg_type = (
                panier_df.groupby("Type EC2")
                .agg(
                    nb_total=("Quantité", "sum"),
                    cout_mois=("$/mois total", "sum"),
                    cout_an=("$/an total", "sum"),
                )
                .reset_index()
                .sort_values("cout_mois", ascending=False)
            )
            fig_type = go.Figure()
            fig_type.add_trace(go.Bar(
                x=agg_type["Type EC2"],
                y=agg_type["cout_mois"],
                name="$/mois",
                marker_color="#636EFA",
                text=[f"${v:,.0f}" for v in agg_type["cout_mois"]],
                textposition="outside",
                hovertemplate="<b>%{x}</b><br>$/mois : $%{y:,.2f}<br>Nb instances : %{customdata}<extra></extra>",
                customdata=agg_type["nb_total"],
            ))
            fig_type.add_trace(go.Bar(
                x=agg_type["Type EC2"],
                y=agg_type["cout_an"],
                name="$/an",
                marker_color="#EF553B",
                text=[f"${v:,.0f}" for v in agg_type["cout_an"]],
                textposition="outside",
                hovertemplate="<b>%{x}</b><br>$/an : $%{y:,.2f}<extra></extra>",
                visible="legendonly",
            ))
            fig_type.update_layout(
                title="Coût mensuel & annuel cumulé par type d'instance EC2",
                xaxis_title="Type EC2",
                yaxis_title="Coût ($)",
                barmode="group",
                height=450,
                hovermode="x unified",
                legend=dict(orientation="h", yanchor="bottom", y=1.02),
            )
            st.plotly_chart(fig_type, width="stretch")

        with tab_h2:
            agg_reg = (
                panier_df.groupby("Région")
                .agg(
                    nb_total=("Quantité", "sum"),
                    cout_mois=("$/mois total", "sum"),
                    cout_an=("$/an total", "sum"),
                )
                .reset_index()
                .sort_values("cout_mois", ascending=False)
            )
            fig_reg = px.bar(
                agg_reg,
                x="Région",
                y="cout_mois",
                color="Région",
                text=agg_reg["cout_mois"].apply(lambda v: f"${v:,.0f}"),
                title="Coût mensuel cumulé par région",
                labels={"cout_mois": "$/mois", "Région": "Région"},
                color_discrete_sequence=px.colors.qualitative.Set2,
            )
            fig_reg.update_traces(textposition="outside")
            fig_reg.update_layout(height=430, showlegend=False)
            st.plotly_chart(fig_reg, width="stretch")

        with tab_h3:
            agg_eng = (
                panier_df.groupby("Engagement")
                .agg(
                    nb_total=("Quantité", "sum"),
                    cout_mois=("$/mois total", "sum"),
                )
                .reset_index()
            )
            eng_labels = {"OD": "On-Demand", "1an": "Reserved 1 an", "3ans": "Reserved 3 ans", "SP": "Savings Plan"}
            agg_eng["Engagement Label"] = agg_eng["Engagement"].map(eng_labels)
            fig_eng = px.pie(
                agg_eng,
                values="cout_mois",
                names="Engagement Label",
                title="Répartition du coût mensuel par type d'engagement",
                hole=0.4,
                color_discrete_sequence=px.colors.qualitative.Pastel,
            )
            fig_eng.update_layout(height=400)
            st.plotly_chart(fig_eng, width="stretch")

        # ── Métriques totales ─────────────────────────────────────────────
        st.markdown("---")
        st.subheader("💰 Totaux du panier")
        tot_inst = int(panier_df["Quantité"].sum())
        tot_mois = panier_df["$/mois total"].sum()
        tot_an   = panier_df["$/an total"].sum()
        tot_cpu  = int((panier_df["vCPU unit."] * panier_df["Quantité"]).sum())
        tot_ram  = (panier_df["RAM unit. (Go)"] * panier_df["Quantité"]).sum()

        m1, m2, m3, m4, m5 = st.columns(5)
        m1.metric("Instances totales", f"{tot_inst:,}")
        m2.metric("Coût mensuel total", f"${tot_mois:,.2f}")
        m3.metric("Coût annuel total",  f"${tot_an:,.2f}")
        m4.metric("vCPU totaux",        f"{tot_cpu:,}")
        m5.metric("RAM totale",         f"{tot_ram:,.0f} Go")

        # ── Tableau détaillé ─────────────────────────────────────────────
        st.markdown("---")
        st.subheader("📋 Détail du panier")
        display_panier = panier_df[[
            "Région", "Type EC2", "OS", "Engagement",
            "Quantité", "vCPU unit.", "RAM unit. (Go)",
            "$/h unit.", "$/h total", "$/mois total", "$/an total",
        ]].copy()
        st.dataframe(
            display_panier.style.format({
                "$/h unit.":    "${:.4f}",
                "$/h total":    "${:.4f}",
                "$/mois total": "${:,.2f}",
                "$/an total":   "${:,.2f}",
            }),
            width="stretch",
            hide_index=True,
        )

        # ── Export Excel panier ───────────────────────────────────────────
        st.markdown("---")
        st.subheader("📥 Exporter le panier")

        def build_panier_excel(detail_df, agg_type_df, agg_reg_df, totaux):
            wb = openpyxl.Workbook()
            HDR_FILL   = PatternFill("solid", fgColor="232F3E")
            HDR_FONT   = Font(bold=True, color="FFFFFF", size=11)
            ORANGE     = PatternFill("solid", fgColor="FF9900")
            ALT        = PatternFill("solid", fgColor="F8F9FA")
            BORDER     = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"),
            )
            CENTER = Alignment(horizontal="center", vertical="center")

            # Feuille 1 — Détail
            ws1 = wb.active
            ws1.title = "Détail du panier"
            ws1.merge_cells("A1:K1")
            h = ws1["A1"]
            h.value = f"🧮  Panier d'instances EC2 — généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
            h.font = Font(bold=True, size=13, color="FFFFFF")
            h.fill = ORANGE
            h.alignment = CENTER
            ws1.row_dimensions[1].height = 26

            cols_d = list(detail_df.columns)
            for ci, hdr in enumerate(cols_d, 1):
                c = ws1.cell(row=3, column=ci, value=hdr)
                c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = BORDER
            ws1.row_dimensions[3].height = 18

            money_cols = {"$/h unit.", "$/h total", "$/mois total", "$/an total"}
            for ri, row_d in detail_df.iterrows():
                er = ri + 4
                fill = ALT if ri % 2 else PatternFill("solid", fgColor="FFFFFF")
                for ci, (col_n, val) in enumerate(zip(cols_d, row_d), 1):
                    c = ws1.cell(row=er, column=ci, value=val)
                    c.border = BORDER; c.alignment = CENTER; c.fill = fill
                    if col_n in money_cols and isinstance(val, float):
                        c.number_format = '#,##0.0000 "$"' if "h" in col_n else '#,##0.00 "$"'
                ws1.row_dimensions[er].height = 15

            # ligne totaux
            tot_row = len(detail_df) + 4
            ws1.cell(row=tot_row, column=1, value="TOTAL").font = Font(bold=True)
            ws1.cell(row=tot_row, column=5, value=totaux["inst"]).font = Font(bold=True)
            ws1.cell(row=tot_row, column=9, value=totaux["h_total"]).number_format = '#,##0.0000 "$"'
            ws1.cell(row=tot_row, column=10, value=totaux["mois"]).number_format = '#,##0.00 "$"'
            ws1.cell(row=tot_row, column=11, value=totaux["an"]).number_format = '#,##0.00 "$"'
            for ci in range(1, 12):
                ws1.cell(row=tot_row, column=ci).fill = PatternFill("solid", fgColor="FFF3CD")
                ws1.cell(row=tot_row, column=ci).border = BORDER
            ws1.row_dimensions[tot_row].height = 17

            col_widths_d = [30, 20, 10, 10, 10, 12, 16, 14, 14, 16, 16]
            for i, w in enumerate(col_widths_d, 1):
                ws1.column_dimensions[get_column_letter(i)].width = w

            # Feuille 2 — Par type
            ws2 = wb.create_sheet("Par type EC2")
            for ci, hdr in enumerate(["Type EC2", "Nb instances", "$/mois", "$/an"], 1):
                c = ws2.cell(row=1, column=ci, value=hdr)
                c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = BORDER
            for ri, row_t in agg_type_df.iterrows():
                fill = ALT if ri % 2 else PatternFill("solid", fgColor="FFFFFF")
                for ci, val in enumerate([row_t["Type EC2"], row_t["nb_total"], row_t["cout_mois"], row_t["cout_an"]], 1):
                    c = ws2.cell(row=ri+2, column=ci, value=val)
                    c.border = BORDER; c.alignment = CENTER; c.fill = fill
                    if ci in (3, 4): c.number_format = '#,##0.00 "$"'
            for i, w in enumerate([22, 14, 16, 16], 1):
                ws2.column_dimensions[get_column_letter(i)].width = w

            # Feuille 3 — Par région
            ws3 = wb.create_sheet("Par région")
            for ci, hdr in enumerate(["Région", "Nb instances", "$/mois", "$/an"], 1):
                c = ws3.cell(row=1, column=ci, value=hdr)
                c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = BORDER
            for ri, row_r in agg_reg_df.iterrows():
                fill = ALT if ri % 2 else PatternFill("solid", fgColor="FFFFFF")
                for ci, val in enumerate([row_r["Région"], row_r["nb_total"], row_r["cout_mois"], row_r["cout_an"]], 1):
                    c = ws3.cell(row=ri+2, column=ci, value=val)
                    c.border = BORDER; c.alignment = CENTER; c.fill = fill
                    if ci in (3, 4): c.number_format = '#,##0.00 "$"'
            for i, w in enumerate([32, 14, 16, 16], 1):
                ws3.column_dimensions[get_column_letter(i)].width = w

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf.read()

        panier_excel = build_panier_excel(
            display_panier,
            agg_type,
            agg_reg,
            {
                "inst":    tot_inst,
                "h_total": panier_df["$/h total"].sum(),
                "mois":    tot_mois,
                "an":      tot_an,
            }
        )
        fname_panier = f"panier_ec2_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        col_ex1, col_ex2 = st.columns([1, 3])
        with col_ex1:
            st.download_button(
                label="📊 Télécharger le panier Excel",
                data=panier_excel,
                file_name=fname_panier,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.caption("3 onglets : Détail · Par type · Par région")
        with col_ex2:
            st.info("💡 Le panier est conservé pendant toute votre session. Utilisez **Vider le panier** pour repartir de zéro.")

    else:
        st.info("Le panier est vide. Utilisez le formulaire ci-dessus pour ajouter des instances.")
